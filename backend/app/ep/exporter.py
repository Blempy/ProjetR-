from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from .schemas import CollectorResult, EPResult


TEMPLATE_PATH = Path("templates/excel/note_ep.xlsx")
EXPORT_ROOT = Path("exports/ep")


def _write_row(ws, row_idx: int, values: Iterable):
    for col_idx, value in enumerate(values, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)


def generate_excel(result: EPResult) -> Path:
    if not TEMPLATE_PATH.exists():
        raise FileNotFoundError(f"Template Excel introuvable : {TEMPLATE_PATH}")

    wb = load_workbook(TEMPLATE_PATH)

    ws_params = wb["Paramètres"]
    ws_params.cell(row=2, column=2, value=result.rain.method)
    if result.collectors:
        ws_params.cell(row=3, column=2, value=result.collectors[0].permitted_outflow_lps)
        ws_params.cell(row=4, column=2, value=result.collectors[0].safety_factor_applied)

    ws_surfaces = wb["Surfaces"]
    ws_collectors = wb["Collecteurs"]
    ws_autocurage = wb["Autocurage"]
    ws_synthese = wb["Synthèse"]

    surface_row = 2
    collector_row = 2
    autocurage_row = 2

    total_volume = 0.0

    for collector in result.collectors:
        # Surfaces: regroupées dans result? (manque surfaces). Placeholder: total par collecteur.
        _write_row(
            ws_surfaces,
            surface_row,
            [
                f"Total {collector.id}",
                "",
                collector.area_total_m2,
                collector.weighted_coefficient,
                collector.id,
            ],
        )
        surface_row += 1

        _write_row(
            ws_collectors,
            collector_row,
            [
                collector.id,
                "-",
                "-",
                "",
                collector.recommended_diameter_mm,
                collector.inflow_lps,
                collector.permitted_outflow_lps,
                collector.storage_volume_m3,
            ],
        )
        collector_row += 1

        _write_row(
            ws_autocurage,
            autocurage_row,
            [
                collector.id,
                collector.recommended_diameter_mm,
                collector.full_flow_lps,
                collector.inflow_lps,
                collector.velocity_ms,
                collector.safety_factor_applied,
                "OK" if collector.conforms_autoclean else "À contrôler",
            ],
        )
        autocurage_row += 1

        total_volume += collector.storage_volume_m3

    _write_row(
        ws_synthese,
        2,
        ["Volume total", total_volume, "Somme des volumes"],
    )

    export_dir = EXPORT_ROOT / result.project.name.replace(" ", "_")
    export_dir.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.utcnow().strftime("%Y%m%d-%H%M%S")
    output_path = export_dir / f"EP_{timestamp}.xlsx"
    wb.save(output_path)
    return output_path
